#!/usr/bin/env python3
"""Validate a converted Molnify Excel app for common issues.

Usage: python molnify_validate.py <converted_file.xlsx>

Checks for:
- Metadata on first sheet with ID set
- ParseAllSheets when needed
- Correct Molnify colors on active sheets
- molnifyIgnore on model/data sheets
- Green cells with formulas referencing other green cells
- Chart/table blue-cell structure
- 3-cell layout (title, value, UI)
- Metadata key validity
- UI option syntax
- Chart type validity
- Action cell structure
- Unsupported Excel functions
- Autofill metadata keys without corresponding named ranges
- Non-ASCII variable names (incompatible with template engines)
- Hyphenated app IDs (require backtick-quoting in SQL)
- Duplicate variable names across inputs/outputs
- showIfVariable referencing undefined variables
- Action blocks missing required fields per type
- Cross-sheet formula references to non-existent sheets
- Duplicate tab names
- Empty output cells (no value or formula)
- String literals over 256 characters in formulas
- Formulas with an odd number of double-quotes (unbalanced string literals)
- Non-formula cell content exceeding Excel's 32,767 character limit
- Autofill SQL using @variable references (not supported)
- Headless mode with JavaScriptAfterLoad/AfterCalc (ignored in headless)
"""

import json
import re
import sys
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Molnify colour constants
# ---------------------------------------------------------------------------

# All colors accepted by the Molnify backend (exact byte matching, no tolerance).
# Each list's first entry is the recommended color for new apps.
# Source: backend colour-matching logic uses exact equality on RGB hex.
ACCEPTED_COLORS = {
    'input': [
        '00B050',  # Standard Excel green (recommended)
        '008000',  # Excel 2011 Mac green
        '00FF00',  # Google Sheets lime green
        '34A853',  # Google Sheets "standard green"
    ],
    'output': [
        'FF0000',  # Standard red (recommended)
        'EE0000',  # Excel standard red variant
    ],
    'chart': [
        '0070C0',  # Standard Excel blue (recommended)
        '3366FF',  # Excel 2011 Mac blue
        '0000FF',  # Google Sheets blue
    ],
    'action': [
        'FFFF00',  # Standard yellow (recommended)
    ],
    'metadata': [
        '7030A0',  # Standard Excel purple (recommended)
        '660066',  # Excel 2011 Mac purple
        '9900FF',  # Google Sheets purple
        'FF00FF',  # Google Sheets magenta
    ],
}

# Flat lookup: hex -> cell type
_COLOR_TO_TYPE = {}
for _cell_type, _hexes in ACCEPTED_COLORS.items():
    for _h in _hexes:
        _COLOR_TO_TYPE[_h] = _cell_type

# Recommended colors (first entry per type) - used for the "near-miss" warning
RECOMMENDED_COLORS = {hexes[0]: ct for ct, hexes in ACCEPTED_COLORS.items()}

# Tolerance used only for the "near-miss" heuristic (not for matching)
_NEAR_MISS_TOLERANCE = 40


# ---------------------------------------------------------------------------
# Action types (public-facing types only)
# ---------------------------------------------------------------------------

KNOWN_ACTION_TYPES = {
    'email', 'http', 'scenariosave', 'generatereport', 'generatefile',
    'addrecord', 'insertrow', 'downloadquery', 'createesign', 'multiple',
    'aiprompt', 'resetinputs',
}

# Required fields per action type.
ACTION_REQUIRED_FIELDS = {
    'email': ['to'],
    'http': ['url'],
    'generatereport': ['template'],
    'generatefile': ['template'],
    'downloadquery': ['query'],
    'createesign': ['template', 'provider'],
    'aiprompt': ['model', 'prompt'],
}


# ---------------------------------------------------------------------------
# Metadata keys (public-facing keys only)
# ---------------------------------------------------------------------------

KNOWN_METADATA_KEYS = {
    # Identity
    'id', 'name', 'author', 'website', 'version', 'description', 'reference',
    # Panel customization
    'inputpaneltitle', 'outputpaneltitle', 'calculatetitle', 'copytitle', 'savetitle',
    # Feature toggles
    'autocalcenabled', 'enabledforcalculate', 'enabledforsave', 'enabledforreset',
    'enabledforprint', 'enabledforupdate', 'enabledforlogout',
    'enabledfordynamictitles', 'cookieconsentpopup', 'headless',
    # User management
    'users', 'superusers', 'managers', 'realms', 'dbusers',
    # Styling
    'css', 'topbannercolor', 'headertextcolor', 'panelheadercolor',
    'buttoncolor', 'buttonactivecolor', 'successbuttoncolor',
    'expandbuttoncolor', 'collapsebuttoncolor', 'outputboxbackgroundcolor',
    'chartdownloadbackgroundcolor', 'toggleactivecolor',
    'sliderminandmaxbackgroundcolor', 'sliderminandmaxtextcolor',
    'slidercurrentvaluebackgroundcolor', 'slidercurrentvaluetextcolor',
    'headerfont', 'bodyfont', 'logourl', 'faviconurl', 'appletouchiconurl',
    'backgroundimageurl', 'template',
    # JavaScript
    'javascript', 'javascriptafterload', 'javascriptaftercalc',
    'additionalcss', 'additionaljavascript',
    # Scenarios
    'scenarioterm', 'realmterm', 'scenarionamevariable', 'scenariosaveprivate',
    'scenariolockoption', 'scenarioshareduserscell', 'recordtablename',
    'recordtableidvariable', 'recordtableaccesscolumn', 'calctimerseconds',
    # Security
    'ipranges', 'tokenauthentication',
    # Tooltips
    'custominfotooltip', 'customreferencetooltip', 'customupdatetooltip',
    'customsavetooltip', 'customprinttooltip', 'customresettooltip',
    'customlogouttooltip', 'customdownloaddatatooltip', 'customdeletetooltip',
    # Advanced
    'headhtml', 'topbannerhidden', 'headerhidden', 'inputpanelsmall', 'inputpanelfixed',
    'panelsfixed', 'outputboxpanelhidden', 'onlyappnameinpagetitle',
    'clearclipboardatreset', 'clearclipboardatcalc', 'customlogouturl',
    'onlyincludesheet',
}

# Prefixed keys - matched by prefix (case-insensitive)
KNOWN_METADATA_PREFIXES = [
    'parseallsheets',
    'autofill.',
    'datatable.',
]


# ---------------------------------------------------------------------------
# DataTable column types - backend rejects parameterized types and LONG*.
# ---------------------------------------------------------------------------

VALID_COLUMN_TYPES = {
    # Text
    'VARCHAR', 'TEXT', 'TINYTEXT', 'MEDIUMTEXT', 'CHAR',
    # Integer (UNSIGNED variants handled separately)
    'INT', 'TINYINT', 'SMALLINT', 'MEDIUMINT', 'BIGINT',
    'INT UNSIGNED', 'TINYINT UNSIGNED', 'SMALLINT UNSIGNED',
    'MEDIUMINT UNSIGNED', 'BIGINT UNSIGNED',
    # Decimal
    'DECIMAL', 'DOUBLE', 'FLOAT', 'REAL', 'NUMERIC',
    # Boolean
    'BOOL',
    # Date/Time
    'DATETIME', 'TIMESTAMP', 'DATE', 'TIME', 'YEAR',
    # Binary
    'BLOB', 'TINYBLOB', 'MEDIUMBLOB', 'VARBINARY', 'BINARY',
    # Other
    'JSON', 'SET', 'ENUM',
}

BLOCKED_TYPE_PREFIXES = ['LONG']


# ---------------------------------------------------------------------------
# UI options
# ---------------------------------------------------------------------------

# Bare tokens (no = sign)
KNOWN_UI_BARE_OPTIONS = {
    'slider', 'range', 'dropdown', 'select', 'multiple', 'date', 'time',
    'textarea', 'nobuttons', 'fileupload', 'signature', 'barcode',
    'infotext', 'longinfotext', 'hidden', 'notitle', 'html',
    'peoplematrix', 'amonginputs', 'hidecopy', 'leftcolumn', 'rightcolumn',
    'panelhidden', 'resetwhenhidden', 'nocalc',
    'recorddropdown', 'rowdropdown',
}

# Key=value tokens (only the key part)
KNOWN_UI_KEYS = {
    'min', 'max', 'delta', 'gridnum', 'multiplier', 'prefix', 'postfix',
    'decimals', 'placeholder', 'regex', 'class', 'dividername', 'tab',
    'showifvariable', 'showifvalue', 'showifvaluenot', 'variable',
    'jsonchange', 'icon', 'background', 'color', 'columns', 'rows',
    'maxfiles', 'acceptedfiletypes',
    'display', 'tablename', 'map', 'idvariable', 'filter', 'orderby',
    'maxrows', 'accesscolumn',
}

# User context variable tokens (bare, case-insensitive)
KNOWN_UI_AUTOFILL = {
    'user', 'user.email', 'user.ip',
    'user.identificationnumber', 'user.authenticationprovider',
    'superuser', 'manager', 'currentrealm',
}
# tag1-tag10
for _i in range(1, 11):
    KNOWN_UI_AUTOFILL.add(f'tag{_i}')


# ---------------------------------------------------------------------------
# Chart types
# ---------------------------------------------------------------------------

KNOWN_CHART_TYPES = {
    'barchart', 'linechart', 'piechart', 'donutchart', 'waterfallchart',
    'scatterchart', 'geochart', 'linebarchart', 'table',
}

# Chart key=value UI options
KNOWN_CHART_KEYS = {'xaxis', 'yaxis', 'yaxisticks', 'axisdecimals', 'atleast', 'map'}

# Chart bare-token UI options
KNOWN_CHART_BARE_OPTIONS = {
    'stacked', 'horizontal', 'showvalues', 'nogridlines', 'steps',
    'staggerlabels', 'nowordwrap', 'nocontrols', 'nogroupedstackedcontrols',
    'atleastsync', 'centerzero', 'hidemaxmin',
}


# ---------------------------------------------------------------------------
# Supported Excel functions
# Source: Molnify's publicly supported function list.
# Molnify extension functions have the " (Molnify extension)" suffix stripped.
# ---------------------------------------------------------------------------

SUPPORTED_FUNCTIONS = {
    'ABS', 'ACOS', 'ACOSH', 'ADDRESS', 'AND', 'AREAS', 'ASIN', 'ASINH',
    'ATAN', 'ATAN2', 'ATANH', 'AVEDEV', 'AVERAGE', 'AVERAGEA', 'AVERAGEIF',
    'AVERAGEIFS', 'BIN2DEC', 'CEILING', 'CEILING.MATH', 'CEILING.PRECISE',
    'CHAR', 'CHOOSE', 'CLEAN', 'CODE', 'COLUMN', 'COLUMNS', 'COMBIN',
    'CONCATENATE', 'CORREL', 'COS', 'COSH', 'COUNT', 'COUNTA', 'COUNTBLANK',
    'COUNTIF', 'COUNTIFS', 'COVAR', 'COVARIANCE.P', 'COVARIANCE.S',
    'CUSTOMSQL', 'DATE', 'DATEVALUE', 'DAVERAGE', 'DAY', 'DAYS', 'DAYS360',
    'DBLOOKUP', 'DCOUNT', 'DCOUNTA', 'DEC2BIN', 'DEC2HEX', 'DEGREES',
    'DELTA', 'DEVSQ', 'DGET', 'DMAX', 'DMIN', 'DOLLAR', 'DOLLARDE',
    'DOLLARFR', 'DPRODUCT', 'DSTDEV', 'DSTDEVP', 'DSUM', 'DVAR', 'DVARP',
    'EDATE', 'EOMONTH', 'ERROR.TYPE', 'EVAL', 'EVEN', 'EXACT', 'EXP', 'FACT',
    'FACTDOUBLE', 'FALSE', 'FILTERJSON', 'FILTERXML', 'FIND', 'FIXED',
    'FLOOR', 'FLOOR.MATH', 'FLOOR.PRECISE', 'FORECAST', 'FORECAST.LINEAR',
    'FREQUENCY', 'FV', 'GCD', 'GEOMEAN', 'HEX2DEC', 'HLOOKUP', 'HOUR',
    'HYPERLINK', 'IF', 'IFERROR', 'IFNA', 'IFS', 'IMAGINARY', 'IMREAL',
    'INDEX', 'INDIRECT', 'INT', 'INTERCEPT', 'IPMT', 'IRR', 'ISBLANK',
    'ISERR', 'ISERROR', 'ISEVEN', 'ISLOGICAL', 'ISNA', 'ISNONTEXT',
    'ISNUMBER', 'ISODD', 'ISREF', 'ISTEXT', 'LARGE', 'LCM', 'LEFT', 'LEN',
    'LN', 'LOG', 'LOG10', 'LOOKUP', 'LOWER', 'MATCH', 'MAX', 'MAXA',
    'MAXIFS', 'MDETERM', 'MEDIAN', 'MID', 'MIN', 'MINA', 'MINIFS', 'MINUTE',
    'MINVERSE', 'MIRR', 'MMULT', 'MOD', 'MODE', 'MONTH', 'MROUND', 'NA',
    'NETWORKDAYS', 'NORM.DIST', 'NORM.INV', 'NORM.S.DIST', 'NORM.S.INV',
    'NORMDIST', 'NORMINV', 'NORMSDIST', 'NORMSINV', 'NOT', 'NOW', 'NPER',
    'NPV', 'NUMBERVALUE', 'OCT2DEC', 'ODD', 'OFFSET', 'OR', 'PEARSON',
    'PERCENTILE', 'PERCENTRANK', 'PERCENTRANK.EXC', 'PERCENTRANK.INC', 'PI',
    'PMT', 'POISSON', 'POISSON.DIST', 'POISSONADVANCED', 'POWER', 'PPMT',
    'PRODUCT', 'PROPER', 'PV', 'QUOTIENT', 'RADIANS', 'RAND', 'RANDBETWEEN',
    'RANK', 'RATE', 'REPLACE', 'REPT', 'RIGHT', 'ROMAN', 'ROUND',
    'ROUNDDOWN', 'ROUNDUP', 'ROW', 'ROWS', 'SEARCH', 'SECOND', 'SIGN',
    'SIN', 'SINGLE', 'SINH', 'SLOPE', 'SMALL', 'SOLVER', 'SOLVERBI',
    'SOLVERSTEP', 'SQRT', 'SQRTPI', 'STANDARDIZE', 'STDEV', 'STDEV.P',
    'STDEV.S', 'STDEVA', 'STDEVP', 'STDEVPA', 'SUBSTITUTE', 'SUBTOTAL',
    'SUM', 'SUMIF', 'SUMIFS', 'SUMPRODUCT', 'SUMSQ', 'SUMX2MY2', 'SUMX2PY2',
    'SUMXMY2', 'SWITCH', 'T', 'T.DIST', 'T.DIST.2T', 'T.DIST.RT', 'TAN',
    'TANH', 'TDIST', 'TEXT', 'TEXTJOIN', 'TIME', 'TIMEVALUE', 'TODAY',
    'TRANSPOSE', 'TREND', 'TRIM', 'TRUE', 'TRUNC', 'UPPER', 'VALUE', 'VAR',
    'VAR.P', 'VAR.S', 'VARA', 'VARP', 'VARPA', 'VLOOKUP', 'WEEKDAY',
    'WEEKNUM', 'WORKDAY', 'WORKDAY.INTL', 'XMLSERVICE', 'YEAR', 'YEARFRAC',
}

# Functions regex - matches function calls like SUM(, VLOOKUP(, CEILING.MATH(
_FUNCTION_CALL_RE = re.compile(r'\b([A-Z][A-Z0-9.]*)\s*\(', re.IGNORECASE)

# String literals in formulas - matches "..." (used by multiple checks)
_STRING_LITERAL_RE = re.compile(r'"([^"]*)"')


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def get_cell_color_hex(cell):
    """Return uppercase 6-char hex color if cell has solid fill, else None."""
    fill = cell.fill
    if fill.fill_type != 'solid':
        return None
    color = fill.fgColor
    if color and color.type == 'rgb' and color.rgb:
        return color.rgb[-6:].upper()
    return None


def color_matches_molnify(hex_color):
    """Check if a hex color is one of the exact colors the backend accepts."""
    if not hex_color:
        return None
    return _COLOR_TO_TYPE.get(hex_color.upper())


def is_near_molnify_color(hex_color):
    """Check if a color is close to a recommended Molnify color but not accepted."""
    if not hex_color:
        return False
    if color_matches_molnify(hex_color):
        return False
    r1, g1, b1 = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    for standard_hex in RECOMMENDED_COLORS:
        r2, g2, b2 = int(standard_hex[0:2], 16), int(standard_hex[2:4], 16), int(standard_hex[4:6], 16)
        if (abs(r1 - r2) < _NEAR_MISS_TOLERANCE
                and abs(g1 - g2) < _NEAR_MISS_TOLERANCE
                and abs(b1 - b2) < _NEAR_MISS_TOLERANCE):
            return True
    return False


def has_molnify_ignore(ws):
    """Check if sheet has molnifyIgnore in cell A1."""
    a1 = ws.cell(row=1, column=1).value
    return a1 and str(a1).strip().lower() == 'molnifyignore'


def parse_cell_refs_in_formula(formula, sheet_name):
    """Extract cell references from a formula. Returns set of (sheet, coord)."""
    if not formula or not str(formula).startswith('='):
        return set()
    refs = set()
    formula_str = str(formula)
    # Sheet-qualified references (openpyxl may store \! instead of !)
    pattern = r"(?:'([^']+)'|([A-Za-z_]\w*))\\?!\$?([A-Z]{1,3})\$?(\d+)"
    for m in re.finditer(pattern, formula_str):
        ref_sheet = m.group(1) or m.group(2)
        refs.add((ref_sheet, f"{m.group(3)}{m.group(4)}"))
    # Plain references
    plain = r"(?<![!\w])\$?([A-Z]{1,3})\$?(\d+)"
    for m in re.finditer(plain, formula_str):
        refs.add((sheet_name, f"{m.group(1)}{m.group(2)}"))
    return refs


def _is_known_metadata_key(key_lower):
    """Check if a metadata key is known (exact match or prefix match)."""
    if key_lower in KNOWN_METADATA_KEYS:
        return True
    for prefix in KNOWN_METADATA_PREFIXES:
        if key_lower == prefix.rstrip('.') or key_lower.startswith(prefix):
            return True
    return False


def _is_known_ui_token(token_lower):
    """Check if a UI token (bare or key=value) is recognized."""
    if '=' in token_lower:
        key = token_lower.split('=', 1)[0]
        # seriesN.property pattern
        if re.match(r'^series\d+\.\w+$', key):
            return True
        return key in KNOWN_UI_KEYS or key in KNOWN_CHART_KEYS
    return (token_lower in KNOWN_UI_BARE_OPTIONS
            or token_lower in KNOWN_UI_AUTOFILL
            or token_lower in KNOWN_CHART_BARE_OPTIONS
            or token_lower.startswith('jsonchange'))


def _iter_metadata_pairs(wb, colored_cells):
    """Yield (sheet, coord, key_str, val_cell) for each purple metadata key cell.

    Skips the value half of a pair (a purple cell whose left neighbor is also
    purple). val_cell is the cell one column to the right of the key.
    """
    for (sheet, coord), ctype in colored_cells.items():
        if ctype != 'metadata':
            continue
        ws = wb[sheet]
        cell = ws[coord]
        if cell.value is None:
            continue
        if cell.column >= 2:
            left_cell = ws.cell(row=cell.row, column=cell.column - 1)
            left_color = get_cell_color_hex(left_cell)
            if left_color and color_matches_molnify(left_color) == 'metadata':
                continue
        val_cell = ws.cell(row=cell.row, column=cell.column + 1)
        yield sheet, coord, str(cell.value).strip(), val_cell


def _iter_ui_strings(wb, colored_cells, types=('input', 'output')):
    """Yield (sheet, coord, ui_cell, ui_str, tokens) for each colored cell whose
    UI cell (one column right) holds a literal option string.

    Skips empty and formula-valued UI cells. tokens is the ;-split, stripped,
    non-empty token list.
    """
    for (sheet, coord), ctype in colored_cells.items():
        if ctype not in types:
            continue
        ws = wb[sheet]
        cell = ws[coord]
        ui_cell = ws.cell(row=cell.row, column=cell.column + 1)
        if ui_cell.value is None:
            continue
        ui_str = str(ui_cell.value).strip()
        if ui_str.startswith('='):
            continue
        tokens = [t.strip() for t in ui_str.split(';') if t.strip()]
        yield sheet, coord, ui_cell, ui_str, tokens


# ---------------------------------------------------------------------------
# Main validation
# ---------------------------------------------------------------------------

def validate_workbook(filepath):
    """Validate a Molnify Excel app and report issues."""
    wb = load_workbook(filepath)
    issues = []

    def error(msg):
        issues.append(('ERROR', msg))

    def warning(msg):
        issues.append(('WARNING', msg))

    def info(msg):
        issues.append(('INFO', msg))

    # Collect colored cells per sheet (only on non-ignored sheets)
    colored_cells = {}  # (sheet, coord) -> cell_type
    green_cells = {}    # (sheet, coord) -> cell value/formula
    sheets_with_colors = set()
    ignored_sheets = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if has_molnify_ignore(ws):
            ignored_sheets.add(sheet_name)
            continue

        for row in ws.iter_rows():
            for cell in row:
                hex_color = get_cell_color_hex(cell)
                if not hex_color:
                    continue

                cell_type = color_matches_molnify(hex_color)
                if cell_type:
                    colored_cells[(sheet_name, cell.coordinate)] = cell_type
                    sheets_with_colors.add(sheet_name)
                    if cell_type == 'input':
                        green_cells[(sheet_name, cell.coordinate)] = cell.value
                elif is_near_molnify_color(hex_color):
                    warning(f"{sheet_name}!{cell.coordinate}: fill color #{hex_color} is close to "
                            f"a Molnify color but may not be recognized. Use exact standard colors.")
                # Non-Molnify colors on active sheets are fine (uncolored/decorative)

    # --- Check 1: Metadata on first sheet ---
    first_sheet = wb.sheetnames[0]
    first_ws = wb[first_sheet]
    metadata_cells = {k: v for k, v in colored_cells.items() if v == 'metadata'}

    if not metadata_cells:
        error("No purple (metadata) cells found in the workbook.")
    else:
        metadata_sheets = {sheet for sheet, _ in metadata_cells}
        if first_sheet not in metadata_sheets:
            error(f"Metadata must be on the first sheet ('{first_sheet}'), "
                  f"but found on: {', '.join(metadata_sheets)}")

    # --- Check 2: ID metadata is set ---
    has_id = False
    app_id = None
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'metadata':
            ws = wb[sheet]
            cell = ws[coord]
            if cell.value and str(cell.value).strip().upper() == 'ID':
                # Check next cell has a value
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if next_cell.value:
                    has_id = True
                    app_id = str(next_cell.value).strip()
    if not has_id:
        error("No 'ID' metadata found. Set an ID for stable URLs and database table access.")
    elif not re.fullmatch(r'[a-z0-9_-]+', app_id):
        invalid = sorted({c for c in app_id if not re.match(r'[a-z0-9_-]', c)})
        if any(c.isupper() for c in app_id):
            error(f"App ID '{app_id}' has uppercase letter(s): "
                  f"{', '.join(repr(c) for c in invalid)}. IDs must be lowercase - the "
                  f"backend lowercases it, so the table name (data_<id>_0) and lab/prod "
                  f"deployment no longer match the ID the app references.")
        else:
            error(f"App ID '{app_id}' has invalid character(s): "
                  f"{', '.join(repr(c) for c in invalid)}. Use only lowercase letters, "
                  f"digits, hyphens (-) and underscores (_).")

    # --- Check 2b: At least one input or output cell ---
    # Molnify treats a workbook with no input (green) or output (red) cells as
    # invalid, even if the app is driven entirely by JavaScript actions.
    has_input = any(ctype == 'input' for ctype in colored_cells.values())
    has_output = any(ctype == 'output' for ctype in colored_cells.values())
    if not has_input and not has_output:
        error("No input (green) or output (red) cells found. Molnify requires at least "
              "one input or output cell; a workbook with none is treated as invalid. "
              "Add an input or output cell, even for JavaScript-driven apps.")

    # --- Check 3: ParseAllSheets ---
    if len(sheets_with_colors) > 1:
        has_parse_all = False
        for (sheet, coord), ctype in colored_cells.items():
            if ctype == 'metadata':
                ws = wb[sheet]
                cell = ws[coord]
                if cell.value and str(cell.value).strip().lower() == 'parseallsheets':
                    next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if next_cell.value and str(next_cell.value).strip().upper() == 'TRUE':
                        has_parse_all = True
        if not has_parse_all:
            error(f"Colored cells found on multiple sheets ({', '.join(sorted(sheets_with_colors))}) "
                  f"but ParseAllSheets is not set to TRUE.")

    # --- Check 4: molnifyIgnore on model sheets ---
    # Sheets that have no colored cells and are not the first sheet (metadata)
    for sheet_name in wb.sheetnames:
        if sheet_name in ignored_sheets:
            continue
        if sheet_name not in sheets_with_colors and sheet_name != first_sheet:
            warning(f"Sheet '{sheet_name}' has no Molnify colored cells and no molnifyIgnore in A1. "
                    f"If this is a model/data sheet, add 'molnifyIgnore' in cell A1 to prevent "
                    f"accidental color interpretation.")

    # --- Check 5: Green cells with formulas referencing other green cells ---
    for (sheet, coord), value in green_cells.items():
        if value and str(value).startswith('='):
            refs = parse_cell_refs_in_formula(str(value), sheet)
            for ref in refs:
                if ref in green_cells:
                    warning(f"{sheet}!{coord}: input (green) cell has formula referencing "
                            f"another input {ref[0]}!{ref[1]}. Input formulas only run once "
                            f"on load - this value won't update when the referenced input changes.")

    # --- Check 6: 3-cell layout (title to the left of colored value cells) ---
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'metadata':
            continue  # Metadata uses 2-cell layout, both purple
        ws = wb[sheet]
        cell = ws[coord]
        if cell.column < 2:
            warning(f"{sheet}!{coord}: {ctype} cell is in column A - no room for a title cell "
                    f"to the left. Molnify expects title in the column before the value cell.")
            continue
        title_cell = ws.cell(row=cell.row, column=cell.column - 1)
        if title_cell.value is None:
            title_color = get_cell_color_hex(title_cell)
            title_type = color_matches_molnify(title_color) if title_color else None
            if not title_type:
                warning(f"{sheet}!{coord}: {ctype} cell has no title in "
                        f"{get_column_letter(cell.column - 1)}{cell.row}. "
                        f"Molnify expects a label in the cell to the left.")

    # --- Check 7: Blue cell chart structure ---
    # Build chart blocks: group blue cells into contiguous row blocks per sheet,
    # tracking the column range of each block.
    blue_cells_by_sheet = {}
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'chart':
            blue_cells_by_sheet.setdefault(sheet, []).append(coord)

    # chart_blocks: list of (sheet, first_data_row, last_data_row, min_col, max_col)
    chart_blocks = []
    for sheet, coords in blue_cells_by_sheet.items():
        ws = wb[sheet]
        # Collect (row, col) for each blue cell
        row_cols = []
        for coord in coords:
            cell = ws[coord]
            row_cols.append((cell.row, cell.column))
        row_cols.sort()

        # Group into contiguous row blocks
        blocks = []
        cur_rows = set()
        cur_cols = set()
        prev_row = None
        for row, col in row_cols:
            if prev_row is not None and row - prev_row > 1:
                blocks.append((sorted(cur_rows), cur_cols.copy()))
                cur_rows = set()
                cur_cols = set()
            cur_rows.add(row)
            cur_cols.add(col)
            prev_row = row
        if cur_rows:
            blocks.append((sorted(cur_rows), cur_cols.copy()))

        for rows, cols in blocks:
            first_row = rows[0]
            last_row = rows[-1]
            min_col = min(cols)
            max_col = max(cols)
            chart_blocks.append((sheet, first_row, last_row, min_col, max_col))

            header_row = first_row - 1
            if header_row < 1:
                warning(f"{sheet}!row {first_row}: chart data starts at row 1 - "
                        f"no room for a header row above.")
                continue

            # Column headers above blue data must not be blue themselves
            for col in sorted(cols):
                header_cell = ws.cell(row=header_row, column=col)
                header_color = get_cell_color_hex(header_cell)
                if header_color and color_matches_molnify(header_color) == 'chart':
                    warning(f"{sheet}!{header_cell.coordinate}: column header cell above "
                            f"blue data is also blue. Headers should have NO color - "
                            f"only data values are blue.")
                    break  # one warning per block is enough

    # --- Check 8: Metadata key validity ---
    for sheet, coord, key_str, _val_cell in _iter_metadata_pairs(wb, colored_cells):
        if not _is_known_metadata_key(key_str.lower()):
            warning(f"{sheet}!{coord}: unrecognized metadata key '{key_str}'. "
                    f"Check for typos - this key will be ignored by the backend.")

    # --- Check 9: UI option syntax ---
    for sheet, coord, ui_cell, ui_str, tokens in _iter_ui_strings(wb, colored_cells):
        for token in tokens:
            token_lower = token.lower()
            if _is_known_ui_token(token_lower):
                continue
            if '=' in token_lower:
                warning(f"{sheet}!{ui_cell.coordinate}: unknown UI option key "
                        f"'{token.split('=', 1)[0]}' in '{ui_str}'.")
            else:
                warning(f"{sheet}!{ui_cell.coordinate}: unknown UI option "
                        f"'{token}' in '{ui_str}'.")

    # --- Check 10: Chart type validity ---
    # Uses chart_blocks from Check 7. For each block, find the UI cell in the
    # header row - it's the rightmost non-empty cell within the block's column
    # range (or one column beyond, where AppBuilder places it).
    for sheet, first_row, last_row, min_col, max_col in chart_blocks:
        ws = wb[sheet]
        header_row = first_row - 1
        if header_row < 1:
            continue

        # Scan the header row for the UI cell: start from one column past the
        # rightmost data column and work outward, then fall back to scanning
        # within the data range from right to left.
        ui_cell = None
        for col_idx in range(max_col + 1, max_col + 4):
            cell = ws.cell(row=header_row, column=col_idx)
            if cell.value is not None:
                ui_cell = cell
                break
        if ui_cell is None:
            # Fall back: rightmost non-empty cell within data range
            for col_idx in range(max_col, min_col - 1, -1):
                cell = ws.cell(row=header_row, column=col_idx)
                if cell.value is not None:
                    ui_cell = cell
                    break

        if ui_cell is None or ui_cell.value is None:
            continue
        ui_str = str(ui_cell.value).strip()
        if ui_str.startswith('='):
            continue
        first_token = ui_str.split(';')[0].strip().lower()
        if first_token and first_token not in KNOWN_CHART_TYPES:
            warning(f"{sheet}!{ui_cell.coordinate}: "
                    f"'{ui_str.split(';')[0].strip()}' is not a recognized "
                    f"chart type. Known types: "
                    f"{', '.join(sorted(KNOWN_CHART_TYPES))}.")

    # --- Check 11: Action cell structure ---
    yellow_cells_by_sheet = {}
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'action':
            yellow_cells_by_sheet.setdefault(sheet, []).append(coord)

    # Collect all action blocks for reuse by Check 21
    all_action_blocks = []  # list of (block_ref, action_kvs)

    for sheet, coords in yellow_cells_by_sheet.items():
        ws = wb[sheet]
        # Group yellow cells into contiguous row blocks
        yellow_rows = {}
        for coord in coords:
            cell = ws[coord]
            yellow_rows.setdefault(cell.row, []).append(cell)

        sorted_row_nums = sorted(yellow_rows.keys())
        blocks = []
        current_block = []
        for i, row_num in enumerate(sorted_row_nums):
            if i == 0 or row_num - sorted_row_nums[i - 1] > 1:
                if current_block:
                    blocks.append(current_block)
                current_block = [row_num]
            else:
                current_block.append(row_num)
        if current_block:
            blocks.append(current_block)

        for block_rows in blocks:
            action_kvs = {}
            block_ref = f"{sheet}!row {block_rows[0]}"
            for row_num in block_rows:
                cells_in_row = sorted(yellow_rows[row_num], key=lambda c: c.column)
                if len(cells_in_row) >= 2:
                    # Both key and value cells are yellow (legacy layout)
                    key_cell = cells_in_row[0]
                    val_cell = cells_in_row[1]
                    if key_cell.value is not None:
                        action_kvs[str(key_cell.value).strip().lower()] = str(
                            val_cell.value) if val_cell.value is not None else ''
                elif len(cells_in_row) == 1:
                    # Only value cell is yellow; key is in the cell to the left
                    val_cell = cells_in_row[0]
                    if val_cell.column >= 2:
                        key_cell = ws.cell(row=row_num, column=val_cell.column - 1)
                        if key_cell.value is not None:
                            action_kvs[str(key_cell.value).strip().lower()] = str(
                                val_cell.value) if val_cell.value is not None else ''

            if not action_kvs:
                continue

            all_action_blocks.append((block_ref, action_kvs))

            if 'type' not in action_kvs:
                warning(f"{block_ref}: action block has no 'type' property. "
                        f"Every action must specify a type.")
            else:
                action_type = action_kvs['type'].strip().lower()
                if action_type not in KNOWN_ACTION_TYPES:
                    warning(f"{block_ref}: unknown action type '{action_kvs['type']}'. "
                            f"Known types: {', '.join(sorted(KNOWN_ACTION_TYPES))}.")

    # --- Check 12: Unsupported Excel functions and _xlfn./_xludf. prefixes ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if has_molnify_ignore(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if not val.startswith('='):
                    continue
                # Check for _xlfn./_xludf. prefixes that Molnify cannot parse
                if '_xlfn.' in val or '_xludf.' in val:
                    prefixes = []
                    for pfx_m in re.finditer(r'(_xlfn\.\w+|_xludf\.\w+)', val):
                        prefixes.append(pfx_m.group(1))
                    error(f"{sheet_name}!{cell.coordinate}: formula contains "
                          f"Excel internal prefix(es) that Molnify cannot "
                          f"parse: {', '.join(prefixes)}. Strip the prefix "
                          f"(e.g. '_xlfn.IFS' -> 'IFS') or rewrite the "
                          f"formula. Functions like IFS, SWITCH, TEXTJOIN, "
                          f"CONCAT, MAXIFS, MINIFS work after prefix removal; "
                          f"XLOOKUP, FILTER, SORT, UNIQUE, LET, LAMBDA are "
                          f"not supported and must be rewritten.")
                # Strip string literals so we don't flag substrings
                # inside quoted text (e.g. "GRADIENT" in an HTML string)
                val_no_strings = re.sub(r'"[^"]*"', '""', val)
                for m in _FUNCTION_CALL_RE.finditer(val_no_strings):
                    func_name = m.group(1).upper()
                    # Skip single-letter matches (e.g. "E" in scientific notation)
                    if len(func_name) <= 1:
                        continue
                    if func_name not in SUPPORTED_FUNCTIONS:
                        warning(f"{sheet_name}!{cell.coordinate}: formula uses "
                                f"'{func_name}' which is not in the list of supported "
                                f"functions. It may cause #NAME? errors at runtime.")

    # --- Check 13: Autofill named ranges ---
    # Collect autofill metadata keys and verify corresponding named ranges exist
    defined_names = {name.lower() for name in wb.defined_names}
    has_autofill = False
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'metadata':
            ws = wb[sheet]
            cell = ws[coord]
            if cell.value is None:
                continue
            key = str(cell.value).strip().lower()
            if key.startswith('autofill.'):
                has_autofill = True
                range_name = key[len('autofill.'):]
                if range_name and range_name not in defined_names:
                    warning(f"{sheet}!{cell.coordinate}: autofill metadata "
                            f"'{cell.value}' references named range "
                            f"'{range_name}', but no such named range is "
                            f"defined in the workbook.")
                # Check for @variable references (not supported in autofill)
                val_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if val_cell.value:
                    sql_text = str(val_cell.value)
                    if sql_text.startswith('='):
                        check_text = ' '.join(
                            m.group(1) for m in _STRING_LITERAL_RE.finditer(sql_text))
                    else:
                        check_text = sql_text
                    at_vars = re.findall(r'@([a-zA-Z]\w*)', check_text)
                    if at_vars:
                        warning(f"{sheet}!{val_cell.coordinate}: autofill SQL "
                                f"contains @variable reference(s): "
                                f"{', '.join('@' + v for v in at_vars)}. "
                                f"Autofill does not support @variable "
                                f"substitution - use a formula to build "
                                f"dynamic SQL instead.")

    # --- Check 14: Non-ASCII variable names ---
    for sheet, coord, ui_cell, ui_str, tokens in _iter_ui_strings(wb, colored_cells):
        for token in tokens:
            if token.lower().startswith('variable='):
                var_name = token.split('=', 1)[1]
                if var_name and not var_name.isascii():
                    warning(f"{sheet}!{ui_cell.coordinate}: variable name "
                            f"'{var_name}' contains non-ASCII characters. "
                            f"Use ASCII-only names to avoid issues with "
                            f"template engines and JavaScript.")

    # --- Check 15: AutoCalcEnabled set to FALSE ---
    for (sheet, coord), ctype in colored_cells.items():
        if ctype == 'metadata':
            ws = wb[sheet]
            cell = ws[coord]
            if cell.value and str(cell.value).strip().lower() == 'autocalcenabled':
                val_cell = ws.cell(row=cell.row, column=cell.column + 1)
                if val_cell.value and str(val_cell.value).strip().upper() == 'FALSE':
                    warning(f"{sheet}!{coord}: AutoCalcEnabled is FALSE - no calculation "
                            f"will run automatically. Avoid unless truly required.")

    # --- Check 16: Hyphenated app ID (info-only, and only when DB features are used) ---
    if app_id and '-' in app_id and has_autofill:
        info(f"App ID '{app_id}' contains hyphens. Table names like "
             f"'data_{app_id}_0' must be backtick-quoted in user-written "
             f"SQL (autofill queries, downloadquery actions).")

    # --- Check 17: DataTable.N schema validation ---
    for sheet, coord, key_str, val_cell in _iter_metadata_pairs(wb, colored_cells):
        if not key_str.lower().startswith('datatable.'):
            continue
        if val_cell.value is None:
            continue
        val_str = str(val_cell.value).strip()
        if not val_str or val_str == '{}':
            continue  # empty or drop-table intent
        try:
            schema = json.loads(val_str)
        except json.JSONDecodeError as e:
            error(f"{sheet}!{val_cell.coordinate}: {key_str} value is not valid JSON: {e}")
            continue
        if not isinstance(schema, dict):
            error(f"{sheet}!{val_cell.coordinate}: {key_str} value must be a JSON object.")
            continue
        columns = schema.get('columns')
        if columns is not None and not isinstance(columns, dict):
            error(f"{sheet}!{val_cell.coordinate}: {key_str} 'columns' must be an object.")
            continue
        if columns:
            for col_name, col_def in columns.items():
                if isinstance(col_def, str):
                    col_type = col_def.strip().upper()
                elif isinstance(col_def, dict) and 'type' in col_def:
                    col_type = str(col_def['type']).strip().upper()
                else:
                    warning(f"{sheet}!{val_cell.coordinate}: {key_str} column "
                            f"'{col_name}' has invalid definition - expected "
                            f"a type string or object with 'type' key.")
                    continue
                # Check for parameterized types
                if '(' in col_type:
                    error(f"{sheet}!{val_cell.coordinate}: {key_str} column "
                          f"'{col_name}' uses parameterized type '{col_type}'. "
                          f"Use bare type names only (e.g. 'DECIMAL' not "
                          f"'DECIMAL(10,2)') - default lengths are applied "
                          f"automatically.")
                    continue
                # Check blocked types
                blocked = False
                for prefix in BLOCKED_TYPE_PREFIXES:
                    if col_type.startswith(prefix):
                        error(f"{sheet}!{val_cell.coordinate}: {key_str} column "
                              f"'{col_name}' uses blocked type '{col_type}'.")
                        blocked = True
                        break
                if blocked:
                    continue
                # Check valid type
                if col_type not in VALID_COLUMN_TYPES:
                    warning(f"{sheet}!{val_cell.coordinate}: {key_str} column "
                            f"'{col_name}' has unrecognized type '{col_type}'. "
                            f"Valid types: VARCHAR, TEXT, INT, DECIMAL, BOOL, "
                            f"DATETIME, TIMESTAMP, JSON, etc.")
                # Check DATETIME/TIMESTAMP default requirement
                if col_type in ('DATETIME', 'TIMESTAMP'):
                    has_default = False
                    if isinstance(col_def, dict) and col_def.get('default'):
                        has_default = True
                    if not has_default:
                        warning(f"{sheet}!{val_cell.coordinate}: {key_str} column "
                                f"'{col_name}' is {col_type} but has no default. "
                                f"DATETIME and TIMESTAMP columns require an "
                                f"explicit default (e.g. 'CURRENT_TIMESTAMP').")

    # --- Check 18: Contrast issues in styling metadata ---
    # Collect metadata key-value pairs for contrast checking
    meta_values = {}
    for sheet, coord, key_str, val_cell in _iter_metadata_pairs(wb, colored_cells):
        if val_cell.value is not None:
            meta_values[key_str.lower()] = (str(val_cell.value).strip(), f"{sheet}!{coord}")

    def _relative_luminance(hex_color):
        """WCAG 2.1 relative luminance from a 6-char hex color string."""
        hex_color = hex_color.lstrip('#')
        if len(hex_color) != 6:
            return None
        try:
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
        except ValueError:
            return None
        channels = []
        for c in (r, g, b):
            s = c / 255.0
            channels.append(s / 12.92 if s <= 0.04045 else ((s + 0.055) / 1.055) ** 2.4)
        return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]

    def _contrast_ratio(hex1, hex2):
        """WCAG contrast ratio between two hex colors. Returns 1.0 on error."""
        l1 = _relative_luminance(hex1)
        l2 = _relative_luminance(hex2)
        if l1 is None or l2 is None:
            return 1.0
        lighter, darker = max(l1, l2), min(l1, l2)
        return (lighter + 0.05) / (darker + 0.05)

    def _white_unreadable_on(bg_hex):
        """Return True if white text on this background fails WCAG AA large text (3:1)."""
        return _contrast_ratio('FFFFFF', bg_hex) < 3.0

    # Elements where the default text is white - a light background makes text invisible
    _WHITE_TEXT_ELEMENTS = {
        'panelheadercolor': 'Panel headers have white text by default',
        'buttoncolor': 'Buttons have white text by default',
        'successbuttoncolor': 'Action buttons have white text by default',
        'outputboxbackgroundcolor': 'Output boxes have white text by default',
        'expandbuttoncolor': 'Expand buttons have white text by default',
        'collapsebuttoncolor': 'Collapse buttons have white text by default',
        'toggleactivecolor': 'Toggle buttons have white text by default',
    }
    for key, reason in _WHITE_TEXT_ELEMENTS.items():
        if key in meta_values:
            color_val, ref = meta_values[key]
            if _white_unreadable_on(color_val):
                warning(f"{ref}: {key} is set to '{color_val}'. "
                        f"{reason} - white text on this background has insufficient "
                        f"contrast (below WCAG 3:1). Use a darker background or "
                        f"override the text color via CSS.")

    # --- Check 19: Variable name uniqueness ---
    defined_variables = {}  # variable_name -> list of (sheet, coord)
    for sheet, coord, ui_cell, ui_str, tokens in _iter_ui_strings(wb, colored_cells):
        for token in tokens:
            if token.lower().startswith('variable='):
                var_name = token.split('=', 1)[1]
                if var_name:
                    defined_variables.setdefault(var_name, []).append(
                        (sheet, coord))
    for var_name, locations in defined_variables.items():
        if len(locations) > 1:
            refs = ', '.join(f"{s}!{c}" for s, c in locations)
            error(f"Variable '{var_name}' is defined {len(locations)} times: "
                  f"{refs}. Each variable name must be unique - the backend "
                  f"returns the first match, making duplicates unreachable.")

    # --- Check 20: ShowIf variable references ---
    known_vars = set(defined_variables.keys())
    for sheet, coord, ui_cell, ui_str, tokens in _iter_ui_strings(wb, colored_cells):
        for token in tokens:
            if token.lower().startswith('showifvariable='):
                ref_var = token.split('=', 1)[1]
                if ref_var and ref_var not in known_vars:
                    error(f"{sheet}!{ui_cell.coordinate}: showIfVariable "
                          f"references '{ref_var}' but no input or output "
                          f"defines variable={ref_var}. Conditional "
                          f"visibility will not work.")
    # Also check action blocks for showIfCell
    for block_ref, action_kvs in all_action_blocks:
        if 'showifvariable' in action_kvs:
            warning(f"{block_ref}: action uses 'showIfVariable' but actions "
                    f"use 'showIfCell' instead. Rename to 'showIfCell'.")
        show_var = action_kvs.get('showifcell')
        if show_var and show_var.strip() and show_var.strip() not in known_vars:
            error(f"{block_ref}: action showIfCell references "
                  f"'{show_var.strip()}' but no input or output defines "
                  f"that variable.")

    # --- Check 21: Action required fields ---
    action_names = {kvs.get('name', '').strip().lower()
                    for _, kvs in all_action_blocks
                    if kvs.get('name', '').strip()}
    for block_ref, action_kvs in all_action_blocks:
        action_type = action_kvs.get('type', '').strip().lower()
        if action_type in ACTION_REQUIRED_FIELDS:
            for field in ACTION_REQUIRED_FIELDS[action_type]:
                if field not in action_kvs or not action_kvs[field].strip():
                    warning(f"{block_ref}: {action_type} action is missing "
                            f"required field '{field}'.")
        if action_type == 'multiple':
            numeric_keys = [k for k in action_kvs if k.isdigit()]
            if not numeric_keys:
                warning(f"{block_ref}: 'multiple' action has no numeric "
                        f"keys (1, 2, ...) referencing sub-actions.")
            else:
                for k in numeric_keys:
                    sub_name = action_kvs[k].strip().lower()
                    if sub_name and sub_name not in action_names:
                        warning(f"{block_ref}: 'multiple' action references "
                                f"sub-action '{action_kvs[k].strip()}' (key "
                                f"{k}) but no action block has name="
                                f"'{action_kvs[k].strip()}'.")

    # --- Check 22: Cross-sheet formula reference resolution ---
    all_sheet_names = {s.lower(): s for s in wb.sheetnames}
    for (sheet, coord), ctype in colored_cells.items():
        ws = wb[sheet]
        cell = ws[coord]
        if cell.value is None or not str(cell.value).startswith('='):
            continue
        refs = parse_cell_refs_in_formula(str(cell.value), sheet)
        for ref_sheet, ref_coord in refs:
            if ref_sheet == sheet:
                continue  # same-sheet reference, skip
            if ref_sheet.lower() not in all_sheet_names:
                error(f"{sheet}!{coord}: formula references sheet "
                      f"'{ref_sheet}' which does not exist in the workbook.")

    # --- Check 23: Duplicate tab names ---
    tab_names = {}  # tab_name_lower -> list of (sheet, coord)
    for sheet, coord, ui_cell, ui_str, tokens in _iter_ui_strings(
            wb, colored_cells, types=('input',)):
        for token in tokens:
            if token.lower().startswith('tab='):
                tab_name = token.split('=', 1)[1]
                if tab_name:
                    tab_names.setdefault(tab_name.lower(), []).append(
                        (tab_name, sheet, coord))
    for tab_lower, occurrences in tab_names.items():
        if len(occurrences) > 1:
            refs = ', '.join(f"{s}!{c}" for _, s, c in occurrences)
            warning(f"Tab name '{occurrences[0][0]}' used {len(occurrences)} "
                    f"times: {refs}. Each tab= creates a separate tab - "
                    f"duplicate names create multiple tabs with the same "
                    f"label.")

    # --- Check 24: Empty output cells ---
    for (sheet, coord), ctype in colored_cells.items():
        if ctype != 'output':
            continue
        ws = wb[sheet]
        cell = ws[coord]
        if cell.value is not None:
            continue
        # Skip if UI indicates html or hidden (populated by JS or intentional)
        ui_cell = ws.cell(row=cell.row, column=cell.column + 1)
        if ui_cell.value:
            ui_lower = str(ui_cell.value).strip().lower()
            if 'html' in ui_lower or 'hidden' in ui_lower:
                continue
        warning(f"{sheet}!{coord}: output (red) cell has no value or formula. "
                f"This will render as an empty output box.")

    # --- Check 25: String literals over 256 characters in formulas ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if has_molnify_ignore(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if not val.startswith('='):
                    continue
                for m in _STRING_LITERAL_RE.finditer(val):
                    if len(m.group(1)) > 256:
                        error(f"{sheet_name}!{cell.coordinate}: formula "
                              f"contains a string literal of "
                              f"{len(m.group(1))} characters (limit 256). "
                              f"Split into shorter segments joined with &.")

    # --- Check 26: Cell content length ---
    # Excel cells have a 32,767 character limit.  Non-formula cells that
    # exceed this are silently truncated, which is particularly common for
    # JavaScript and CSS metadata in headless apps.
    _CELL_CHAR_LIMIT = 32767
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if has_molnify_ignore(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val.startswith('='):
                    continue  # formula length is rarely the issue
                length = len(val)
                if length >= _CELL_CHAR_LIMIT:
                    label = ''
                    hex_color = get_cell_color_hex(cell)
                    if hex_color and color_matches_molnify(hex_color) == 'metadata':
                        if cell.column >= 2:
                            key_cell = ws.cell(row=cell.row,
                                               column=cell.column - 1)
                            if key_cell.value:
                                label = f" (metadata: {key_cell.value})"
                    error(f"{sheet_name}!{cell.coordinate}: cell content "
                          f"is {length:,} characters, exceeding Excel's "
                          f"32,767 character limit{label}. The value "
                          f"will be truncated.")

    # --- Check 27: Headless mode incompatibilities ---
    headless_val = meta_values.get('headless')
    if headless_val and headless_val[0].upper() == 'TRUE':
        for key, label in (('javascriptafterload', 'JavaScriptAfterLoad'),
                           ('javascriptaftercalc', 'JavaScriptAfterCalc')):
            if key in meta_values:
                _, ref = meta_values[key]
                warning(f"{ref}: {label} is set but Headless is TRUE. "
                        f"Only 'JavaScript' metadata runs in headless mode "
                        f"- {label} is ignored. Move the code to "
                        f"'JavaScript'.")

    # --- Check 28: Unbalanced double-quotes in formulas ---
    # A well-formed formula has an even quote count ("" escapes keep parity
    # even). An odd count means an unterminated string literal, which can make
    # Excel strip the sheet's formulas on load.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if has_molnify_ignore(ws):
            continue
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if not val.startswith('='):
                    continue
                q = val.count('"')
                if q % 2 == 1:
                    error(f"{sheet_name}!{cell.coordinate}: formula has an "
                          f"odd number of double-quotes ({q}), so a string "
                          f"literal is unterminated or a stray quote crept in "
                          f"(commonly from concatenating formula text in "
                          f"code). This can make Excel strip the sheet's "
                          f"formulas on load. Verify the quotes balance.")

    # --- Print results ---
    print(f"=== Validation: {filepath} ===")
    print(f"Sheets: {', '.join(wb.sheetnames)}")
    print(f"Ignored (molnifyIgnore): {', '.join(sorted(ignored_sheets)) or 'none'}")
    print(f"Sheets with Molnify colors: {', '.join(sorted(sheets_with_colors)) or 'none'}")
    print()

    errors = [i for i in issues if i[0] == 'ERROR']
    warnings = [i for i in issues if i[0] == 'WARNING']
    infos = [i for i in issues if i[0] == 'INFO']

    if not errors and not warnings:
        print("No issues found.")
    else:
        if errors:
            print(f"ERRORS ({len(errors)}):")
            for _, msg in errors:
                print(f"  ERROR: {msg}")
            print()
        if warnings:
            print(f"WARNINGS ({len(warnings)}):")
            for _, msg in warnings:
                print(f"  WARNING: {msg}")
            print()

    if infos:
        for _, msg in infos:
            print(f"  INFO: {msg}")
        print()

    print(f"Result: {len(errors)} error(s), {len(warnings)} warning(s)")
    return len(errors) == 0


def main():
    if len(sys.argv) != 2:
        print(f"Usage: python {sys.argv[0]} <converted_file.xlsx>")
        return 1
    ok = validate_workbook(sys.argv[1])
    return 0 if ok else 1


if __name__ == '__main__':
    sys.exit(main())
