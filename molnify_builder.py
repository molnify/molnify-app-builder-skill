#!/usr/bin/env python3
"""Programmatic builder for Molnify Excel apps.

Usage:
    from molnify_builder import AppBuilder

    app = AppBuilder("my-app-id", "My App Name")

    # add_input/add_output return the row number for use in formulas
    row_age = app.add_input("Age", 25, ui="slider;min=0;max=120", tooltip="Enter age")
    row_weight = app.add_input("Weight (kg)", 70)
    row_height = app.add_input("Height (cm)", 170)

    # Dropdown with options - creates data validation automatically
    row_gender = app.add_input("Gender", "Male", options=["Male", "Female", "Other"])

    # Use returned rows in formulas - no manual counting
    app.add_output("BMI", f"=App!B{row_weight}/(App!B{row_height}/100)^2", ui="decimals=2")

    # Interleave an output among inputs
    app.add_output("Summary", '="Age: "&App!B1', ui="html;hideCopy", among_inputs=True)

    app.add_chart("Sales", "barChart", series={"Revenue": [100, 200]}, labels=["Q1", "Q2"])
    app.add_action({"type": "email", "to": "user@example.com", "subject": "Results"})
    app.add_metadata("EnabledForSave", "TRUE")
    app.add_model_sheet("Model")
    app.set_cell("Model!A2", "Monthly Rate")
    app.set_cell("Model!B2", "=App!B3/12")
    app.add_named_range("items", "Autofill", "A2:C20")
    app.save("output.xlsx")
"""

import re
import zipfile
import io
from xml.etree import ElementTree

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# OpenXML namespaces
_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_NS_R = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
_NS_CT = 'http://schemas.openxmlformats.org/package/2006/content-types'

# Standard Molnify cell fills - first (recommended) color per type
_FILLS = {
    'input': PatternFill(start_color='00B050', end_color='00B050', fill_type='solid'),
    'output': PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid'),
    'chart': PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid'),
    'action': PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid'),
    'metadata': PatternFill(start_color='7030A0', end_color='7030A0', fill_type='solid'),
}

_WHITE_FONT = Font(color='FFFFFF')

_CELL_ADDR_RE = re.compile(
    r"^(?:(?:'([^']+)'|([A-Za-z_]\w*))!)?"  # optional sheet name
    r"([A-Z]{1,3})(\d+)$"                    # column + row
)

# openpyxl silently truncates any cell string over 32,767 chars the moment it is
# assigned to a cell (not at save), destroying the tail - commonly the end of large
# JS/CSS metadata in headless apps. The guard must therefore run on the raw string
# BEFORE it reaches a cell; once assigned, cell.value is already truncated and no
# check can recover it. Exactly 32,767 is legal, so the guard is strictly '>'.
# (Raw-openpyxl callers that assign cells directly bypass this - molnify_validate.py
# flags a cell sitting at exactly 32,767 as the post-hoc backstop for that path.)
_CELL_CHAR_LIMIT = 32767


def _check_cell_length(value, address=''):
    """Raise if a non-formula string value would be truncated on assignment."""
    if (isinstance(value, str) and not value.startswith('=')
            and len(value) > _CELL_CHAR_LIMIT):
        where = f" ({address})" if address else ''
        raise ValueError(
            f"Cell value{where} is {len(value):,} chars, over Excel's "
            f"{_CELL_CHAR_LIMIT:,} limit - it would be silently truncated. "
            f"Split the content across multiple cells.")


def _has_section_break(ui):
    """Check if a UI string starts a new tab or divider section."""
    if not ui:
        return False
    ui_lower = ui.lower()
    return 'tab=' in ui_lower or 'dividername=' in ui_lower


def _convert_inline_strings(filepath):
    """Convert openpyxl inline strings to shared strings for Molnify compatibility.

    openpyxl writes string cells as t="inlineStr" with <is><t> elements.
    Molnify's calculation engine doesn't propagate changes correctly
    through inline-string cells. This rewrites the xlsx to use the shared
    string table (t="s") instead, and removes empty <v/> elements from
    formula cells.
    """
    ElementTree.register_namespace('', _NS)
    ElementTree.register_namespace('r', _NS_R)

    _REL_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    _SS_REL_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings'
    _SS_CONTENT_TYPE = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml'
    ElementTree.register_namespace('', _REL_NS)  # for rels files

    buf = io.BytesIO()
    shared_strings = []
    ssi = {}  # string -> index

    # First pass: read all entries, collect sheet files and deferred entries
    with zipfile.ZipFile(filepath, 'r') as zin:
        all_items = zin.infolist()
        raw_data = {}
        for item in all_items:
            raw_data[item.filename] = zin.read(item.filename)

    sheet_files = [f for f in raw_data if f.startswith('xl/worksheets/sheet') and f.endswith('.xml')]
    has_shared_strings = 'xl/sharedStrings.xml' in raw_data

    # Parse existing shared strings if present
    if has_shared_strings:
        root = ElementTree.fromstring(raw_data['xl/sharedStrings.xml'])
        for si in root.findall(f'{{{_NS}}}si'):
            t_el = si.find(f'{{{_NS}}}t')
            text = t_el.text if t_el is not None else ''
            if text not in ssi:
                ssi[text] = len(shared_strings)
                shared_strings.append(text)

    # Process sheet files: convert inline strings, remove empty <v/> on formulas
    modified = False
    for filename in sheet_files:
        root = ElementTree.fromstring(raw_data[filename])
        sheet_modified = False

        for c_el in root.iter(f'{{{_NS}}}c'):
            if c_el.get('t') == 'inlineStr':
                is_el = c_el.find(f'{{{_NS}}}is')
                if is_el is not None:
                    t_el = is_el.find(f'{{{_NS}}}t')
                    text = t_el.text if t_el is not None else ''
                    if text not in ssi:
                        ssi[text] = len(shared_strings)
                        shared_strings.append(text)
                    c_el.remove(is_el)
                    v_el = ElementTree.SubElement(c_el, f'{{{_NS}}}v')
                    v_el.text = str(ssi[text])
                    c_el.set('t', 's')
                    sheet_modified = True

            f_el = c_el.find(f'{{{_NS}}}f')
            if f_el is not None:
                v_el = c_el.find(f'{{{_NS}}}v')
                if v_el is not None and not v_el.text:
                    c_el.remove(v_el)
                    sheet_modified = True

        if sheet_modified:
            modified = True
            raw_data[filename] = ElementTree.tostring(root, xml_declaration=True, encoding='UTF-8')

    if not modified and not shared_strings:
        return  # Nothing to do

    # Build shared strings table
    if shared_strings:
        sst = ElementTree.Element(f'{{{_NS}}}sst')
        sst.set('count', str(len(shared_strings)))
        sst.set('uniqueCount', str(len(shared_strings)))
        for text in shared_strings:
            si = ElementTree.SubElement(sst, f'{{{_NS}}}si')
            t_el = ElementTree.SubElement(si, f'{{{_NS}}}t')
            if text and (text[0] in (' ', '\t', '\n') or text[-1] in (' ', '\t', '\n')):
                t_el.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
            t_el.text = text
        raw_data['xl/sharedStrings.xml'] = ElementTree.tostring(sst, xml_declaration=True, encoding='UTF-8')

    # If shared strings file is new, update content types and workbook rels
    if not has_shared_strings and shared_strings:
        # Update [Content_Types].xml
        if '[Content_Types].xml' in raw_data:
            ElementTree.register_namespace('', _NS_CT)
            ct_root = ElementTree.fromstring(raw_data['[Content_Types].xml'])
            has_override = any(
                el.get('PartName') == '/xl/sharedStrings.xml'
                for el in ct_root.findall(f'{{{_NS_CT}}}Override')
            )
            if not has_override:
                override = ElementTree.SubElement(ct_root, f'{{{_NS_CT}}}Override')
                override.set('PartName', '/xl/sharedStrings.xml')
                override.set('ContentType', _SS_CONTENT_TYPE)
            raw_data['[Content_Types].xml'] = ElementTree.tostring(ct_root, xml_declaration=True, encoding='UTF-8')

        # Update xl/_rels/workbook.xml.rels
        rels_file = 'xl/_rels/workbook.xml.rels'
        if rels_file in raw_data:
            ElementTree.register_namespace('', _REL_NS)
            rels_root = ElementTree.fromstring(raw_data[rels_file])
            has_rel = any(
                el.get('Type') == _SS_REL_TYPE
                for el in rels_root.findall(f'{{{_REL_NS}}}Relationship')
            )
            if not has_rel:
                max_id = 0
                for el in rels_root.findall(f'{{{_REL_NS}}}Relationship'):
                    rid = el.get('Id', '')
                    if rid.startswith('rId'):
                        try:
                            max_id = max(max_id, int(rid[3:]))
                        except ValueError:
                            pass
                rel = ElementTree.SubElement(rels_root, f'{{{_REL_NS}}}Relationship')
                rel.set('Id', f'rId{max_id + 1}')
                rel.set('Type', _SS_REL_TYPE)
                rel.set('Target', 'sharedStrings.xml')
            raw_data[rels_file] = ElementTree.tostring(rels_root, xml_declaration=True, encoding='UTF-8')

    # Write everything back
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
        for filename, data in raw_data.items():
            zout.writestr(filename, data)

    with open(filepath, 'wb') as f:
        f.write(buf.getvalue())


class AppBuilder:
    """Build a Molnify app spreadsheet programmatically.

    Creates a workbook with a metadata sheet (always first) and an interface
    sheet. Tracks row positions automatically so you never hardcode cell refs.

    All add_input/add_output methods return the row number where the cell was
    placed. Use this to build formulas:

        row_age = app.add_input("Age", 25)
        row_bmi = app.add_output("BMI", f"=App!B{row_age}/(App!B{row_age+1}/100)^2")
    """

    def __init__(self, app_id, app_name):
        self._app_id = app_id
        self._app_name = app_name
        self._metadata = [
            ('ID', app_id),
            ('Name', app_name),
        ]
        self._items = []  # unified list of (type, data) in insertion order
        self._charts = []
        self._actions = []
        self._model_sheets = []
        self._extra_cells = []  # (sheet, col_str, row_int, value)
        self._named_ranges = []  # (name, sheet, range_str)
        self._data_validations = []  # (row, options_list)
        # Row tracking: computed during save(), but predicted during add_*()
        self._next_row = 1

    def add_input(self, title, value, ui=None, tooltip=None, options=None):
        """Add an input (green cell) to the app.

        Args:
            title: Label shown to the user.
            value: Default value or formula (str starting with '=').
            ui: Semicolon-separated UI options (e.g. "slider;min=0;max=100").
            tooltip: Tooltip text (rendered as Excel cell comment).
            options: List of dropdown options (e.g. ["Yes", "No"]). Automatically
                creates a named range with data validation and adds "dropdown" to
                the UI options.

        Returns:
            The row number on the App sheet where this input will be placed.
        """
        # Insert separator row before a new tab/divider section
        if _has_section_break(ui) and self._items:
            self._next_row += 1
        row = self._next_row
        item = {
            'title': title,
            'value': value,
            'ui': ui,
            'tooltip': tooltip,
        }
        if options is not None:
            item['options'] = list(options)
        self._items.append(('input', item))
        self._next_row += 1
        return row

    def add_output(self, title, value, ui=None, tooltip=None, among_inputs=False):
        """Add an output (red cell) to the app.

        Args:
            title: Label shown to the user.
            value: Formula string (e.g. "=B2*100") or static value.
            ui: Semicolon-separated UI options (e.g. "decimals=2;icon=fa-chart-line").
            tooltip: Tooltip text (rendered as Excel cell comment).
            among_inputs: If True, place this output at the current position
                interleaved with inputs (adds "amongInputs" to UI automatically).
                If False (default), output is placed after all inputs.

        Returns:
            The row number on the App sheet where this output will be placed.
            When among_inputs=False, this is a predicted row that assumes no
            more inputs will be added after this call.
        """
        item = {
            'title': title,
            'value': value,
            'ui': ui,
            'tooltip': tooltip,
        }
        if among_inputs:
            # Interleave: insert at current position in the unified list
            if item.get('ui'):
                if 'amonginputs' not in item['ui'].lower():
                    item['ui'] += ';amongInputs'
            else:
                item['ui'] = 'amongInputs'
            # Insert separator row before a new tab/divider section
            if _has_section_break(ui) and self._items:
                self._next_row += 1
            row = self._next_row
            self._items.append(('output', item))
            self._next_row += 1
            return row
        else:
            # Deferred: will be placed after all inputs + interleaved outputs
            item['_deferred'] = True
            self._items.append(('output', item))
            # Predict row: count how many items are in the list so far,
            # plus the separator row
            return self._predict_deferred_output_row()

    def _predict_deferred_output_row(self):
        """Predict the row number for the next deferred output.

        Deferred outputs are placed after all inline items (inputs and
        among_inputs outputs), with a blank separator row in between.
        Also accounts for separator rows before tab/divider sections.
        """
        inline_items = [(typ, item) for typ, item in self._items
                        if not item.get('_deferred')]
        inline_count = len(inline_items)
        # Count separator rows inserted before tab/divider sections
        section_breaks = sum(
            1 for i, (typ, item) in enumerate(inline_items)
            if i > 0 and _has_section_break(item.get('ui') or '')
        )
        deferred_before = sum(
            1 for typ, item in self._items
            if item.get('_deferred') and item is not self._items[-1][1]
        )
        # Row = inline items + section breaks + separator row + deferred before + 1
        has_inline = inline_count > 0
        return (inline_count + section_breaks
                + (1 if has_inline else 0) + deferred_before + 1)

    def add_chart(self, title, chart_type, series, labels, ui_extra=None):
        """Add a chart (blue cells) to the app.

        Args:
            title: Chart title.
            chart_type: Chart type string (e.g. "barChart", "lineChart").
            series: Dict of {series_name: [values...]} for each data series.
            labels: List of category/row labels.
            ui_extra: Extra chart UI options (appended after chart_type with ;).
        """
        self._charts.append({
            'title': title,
            'chart_type': chart_type,
            'series': series,
            'labels': labels,
            'ui_extra': ui_extra,
        })

    def add_action(self, properties):
        """Add an action (yellow cells) to the app.

        Args:
            properties: Dict of action key-value pairs.
                Must include 'type'. Example:
                {"type": "email", "to": "a@b.com", "subject": "Hi"}
        """
        self._actions.append(dict(properties))

    def add_metadata(self, key, value):
        """Add a metadata entry (purple cells).

        Args:
            key: Metadata key (e.g. "EnabledForSave").
            value: Metadata value (e.g. "TRUE").
        """
        _check_cell_length(value, f"metadata:{key}")
        self._metadata.append((key, value))

    def add_model_sheet(self, name):
        """Register a model/data sheet (gets molnifyIgnore in A1 automatically).

        Use set_cell() to populate it:
            app.add_model_sheet("Model")
            app.set_cell("Model!A2", "Rate")
            app.set_cell("Model!B2", "=App!B3/12")

        Args:
            name: Sheet name.
        """
        self._model_sheets.append(name)

    def add_named_range(self, name, sheet, cell_range):
        """Define a named range in the workbook.

        Named ranges are used for autofill (database queries), DOCX/XLSX
        report templates, and data validation lists.

        Args:
            name: Range name (e.g. "orderItems"). Must be ASCII.
            sheet: Sheet name where the range lives (e.g. "Autofill").
            cell_range: Cell range string (e.g. "A1:C10").

        Example:
            app.add_model_sheet("Autofill")
            app.add_named_range("recentOrders", "Autofill", "A2:D20")
            app.add_metadata("autofill.recentOrders",
                "SELECT name, amount, date, status FROM `data_my-app_0`")
        """
        self._named_ranges.append((name, sheet, cell_range))

    def set_cell(self, address, value):
        """Set a cell's value or formula on any sheet.

        Args:
            address: Cell address, optionally with sheet name.
                Examples: "Model!B2", "'My Sheet'!A1", "B2" (defaults to App sheet).
            value: Cell value - string, number, or formula (str starting with '=').

        Raises:
            ValueError: If the address cannot be parsed.
        """
        m = _CELL_ADDR_RE.match(address)
        if not m:
            raise ValueError(f"Cannot parse cell address: '{address}'")
        _check_cell_length(value, address)
        sheet = m.group(1) or m.group(2) or 'App'
        col_str = m.group(3)
        row_num = int(m.group(4))
        self._extra_cells.append((sheet, col_str, row_num, value))

    def save(self, filepath):
        """Write the app to an Excel file.

        Creates sheets in order: Metadata, App, then any model sheets.
        Automatically adds ParseAllSheets: TRUE (always needed since Metadata and App are separate sheets).

        Args:
            filepath: Output file path (.xlsx).
        """
        wb = Workbook()

        # --- Metadata sheet ---
        meta_ws = wb.active
        meta_ws.title = 'Metadata'

        # Auto-add ParseAllSheets - always needed since Metadata and App are separate sheets
        has_parse_all = any(k == 'ParseAllSheets' for k, _ in self._metadata)
        if not has_parse_all:
            self._metadata.append(('ParseAllSheets', 'TRUE'))

        for row_idx, (key, value) in enumerate(self._metadata, start=1):
            key_cell = meta_ws.cell(row=row_idx, column=1, value=key)
            val_cell = meta_ws.cell(row=row_idx, column=2, value=value)
            key_cell.fill = _FILLS['metadata']
            val_cell.fill = _FILLS['metadata']
            key_cell.font = _WHITE_FONT
            val_cell.font = _WHITE_FONT

        # --- App (interface) sheet ---
        app_ws = wb.create_sheet('App')
        current_row = 1

        # Split items into inline (inputs + among_inputs outputs) and deferred outputs
        inline_items = [(typ, item) for typ, item in self._items if not item.get('_deferred')]
        deferred_outputs = [item for typ, item in self._items if item.get('_deferred')]

        # Write inline items (inputs and interleaved outputs) in order
        for i, (typ, item) in enumerate(inline_items):
            # Separator row before a new tab/divider section
            if i > 0 and _has_section_break(item.get('ui') or ''):
                current_row += 1
            current_row = self._write_cell_row(app_ws, current_row, item, typ)

        # Blank separator row between inline items and deferred outputs
        if inline_items and (deferred_outputs or self._charts):
            current_row += 1

        # Write deferred outputs
        for item in deferred_outputs:
            current_row = self._write_cell_row(app_ws, current_row, item, 'output')

        # Blank separator before charts
        if self._charts and (inline_items or deferred_outputs):
            current_row += 1

        # Charts
        for chart in self._charts:
            current_row = self._write_chart(app_ws, current_row, chart)
            current_row += 1  # blank row between chart blocks

        # Blank separator before actions
        if self._actions and (self._items or self._charts):
            current_row += 1

        # Actions
        for i, action in enumerate(self._actions):
            current_row = self._write_action(app_ws, current_row, action)
            if i < len(self._actions) - 1:
                current_row += 1  # blank row between action blocks

        # --- Model sheets ---
        for name in self._model_sheets:
            ws = wb.create_sheet(name)
            ws.cell(row=1, column=1, value='molnifyIgnore')

        # --- Extra cells from set_cell() ---
        for sheet_name, col_str, row_num, value in self._extra_cells:
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]
            ws[f"{col_str}{row_num}"] = value

        # --- Named ranges ---
        from openpyxl.workbook.defined_name import DefinedName
        for name, sheet, cell_range in self._named_ranges:
            ref = f"'{sheet}'!{cell_range}" if ' ' in sheet else f"{sheet}!{cell_range}"
            dn = DefinedName(name, attr_text=ref)
            wb.defined_names.add(dn)

        # --- Data validations (dropdown options) ---
        if self._data_validations:
            from openpyxl.worksheet.datavalidation import DataValidation
            # Create a hidden sheet for option lists
            opts_sheet_name = '_Options'
            opts_ws = wb.create_sheet(opts_sheet_name)
            opts_ws.cell(row=1, column=1, value='molnifyIgnore')

            for col_idx, (target_row, options) in enumerate(self._data_validations, start=1):
                col_letter = get_column_letter(col_idx)
                for opt_row, opt_val in enumerate(options, start=2):
                    opts_ws.cell(row=opt_row, column=col_idx, value=opt_val)

                # Define named range for this option list
                range_name = f"_opts_{col_idx}"
                last_row = len(options) + 1
                range_ref = f"'{opts_sheet_name}'!${col_letter}$2:${col_letter}${last_row}"
                dn = DefinedName(range_name, attr_text=range_ref)
                wb.defined_names.add(dn)

                # Add data validation to the target cell on App sheet
                dv = DataValidation(type="list", formula1=f"={range_name}")
                dv.sqref = f"B{target_row}"
                app_ws.add_data_validation(dv)

        wb.save(filepath)
        _convert_inline_strings(filepath)

    def _write_cell_row(self, ws, row, item, cell_type):
        """Write a 3-cell row (title, value, UI) and return the next row."""
        # Column A: title (no fill)
        ws.cell(row=row, column=1, value=item['title'])

        # Column B: value (colored)
        val_cell = ws.cell(row=row, column=2, value=item['value'])
        val_cell.fill = _FILLS[cell_type]

        if item.get('tooltip'):
            val_cell.comment = Comment(item['tooltip'], 'AppBuilder')

        # Handle dropdown options: create data validation and auto-add "dropdown" to UI
        ui = item.get('ui') or ''
        if item.get('options'):
            self._data_validations.append((row, item['options']))
            if 'dropdown' not in ui.lower() and 'select' not in ui.lower():
                ui = ('dropdown;' + ui) if ui else 'dropdown'

        # Column C: UI options (no fill)
        if ui:
            ws.cell(row=row, column=3, value=ui)

        return row + 1

    def _write_chart(self, ws, row, chart):
        """Write a chart block and return the next row after the block."""
        series_names = list(chart['series'].keys())
        labels = chart['labels']

        # Header row: title, series names, UI cell
        ws.cell(row=row, column=1, value=chart['title'])
        for col_idx, name in enumerate(series_names, start=2):
            ws.cell(row=row, column=col_idx, value=name)

        # UI cell in rightmost position of header row
        ui_col = len(series_names) + 2
        ui_str = chart['chart_type']
        if chart.get('ui_extra'):
            ui_str += ';' + chart['ui_extra']
        ws.cell(row=row, column=ui_col, value=ui_str)

        row += 1

        # Data rows: label (no color), values (blue)
        for label_idx, label in enumerate(labels):
            ws.cell(row=row, column=1, value=label)
            for col_idx, series_name in enumerate(series_names, start=2):
                values = chart['series'][series_name]
                val_cell = ws.cell(row=row, column=col_idx,
                                   value=values[label_idx] if label_idx < len(values) else '')
                val_cell.fill = _FILLS['chart']
                val_cell.font = _WHITE_FONT
            row += 1

        return row

    def _write_action(self, ws, row, action):
        """Write an action block (key + yellow value cells) and return next row."""
        for key, value in action.items():
            ws.cell(row=row, column=1, value=key)
            val_cell = ws.cell(row=row, column=2, value=value)
            val_cell.fill = _FILLS['action']
            row += 1
        return row
