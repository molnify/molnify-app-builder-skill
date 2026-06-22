#!/usr/bin/env python3
"""Inspect an Excel file to plan its conversion into a Molnify app.

Outputs a structured JSON analysis (default) or human-readable text (--text).

JSON structure:
  {file, sheets, sheet_details: {name: {dimensions, cells, conditional_formatting,
   data_validations, charts}}, named_ranges, dependency_analysis: {potential_inputs,
   potential_outputs, intermediates}}
"""

import argparse
import json
import re
import sys
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def get_cell_color(cell):
    """Return hex color string if cell has a solid fill, else None."""
    fill = cell.fill
    if fill.fill_type != 'solid':
        return None
    color = fill.fgColor
    if color and color.type == 'rgb' and color.rgb:
        return f"#{color.rgb[-6:]}"
    return None


_DYNAMIC_REF_PATTERN = re.compile(r'\b(INDIRECT|OFFSET)\s*\(', re.IGNORECASE)


def detect_dynamic_refs(formula):
    """Detect INDIRECT() and OFFSET() calls in a formula.

    These functions create references that cannot be statically resolved,
    so dependency analysis will be incomplete for cells using them.

    Returns a list of function names found (e.g. ["INDIRECT", "OFFSET"]).
    """
    if not formula or not str(formula).startswith('='):
        return []
    return [m.group(1).upper() for m in _DYNAMIC_REF_PATTERN.finditer(str(formula))]


def parse_cell_refs(formula, sheet_name):
    """Extract cell references from a formula string.

    Returns a set of (sheet, cell_coordinate) tuples.
    Handles: A1, $A$1, Sheet1!A1, 'Sheet Name'!A1, A1:B5 ranges.

    Known limitations:
    - String literals are not excluded - references inside quoted strings
      (e.g. ="Sheet1!A1") will be incorrectly captured.
    - Structured table references (Table1[Column]) are not parsed.
    - INDIRECT() and OFFSET() produce dynamic references that cannot be
      resolved statically. Use detect_dynamic_refs() to flag these.
    """
    if not formula or not str(formula).startswith('='):
        return set()

    refs = set()
    formula_str = str(formula)

    # Match references like Sheet1!A1, 'Sheet Name'!A1:B5, or plain A1:B5 / A1
    pattern = r"(?:'([^']+)'|([A-Za-z_]\w*))!\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?"
    for m in re.finditer(pattern, formula_str):
        ref_sheet = m.group(1) or m.group(2)
        col_start, row_start = m.group(3), int(m.group(4))
        if m.group(5) and m.group(6):
            col_end, row_end = m.group(5), int(m.group(6))
            for r in range(row_start, row_end + 1):
                for c_idx in range(_col_to_num(col_start), _col_to_num(col_end) + 1):
                    refs.add((ref_sheet, f"{get_column_letter(c_idx)}{r}"))
        else:
            refs.add((ref_sheet, f"{m.group(3)}{row_start}"))

    # Match plain cell references (no sheet prefix) like A1, $B$5, A1:C3
    # Negative lookbehind to avoid matching already-captured sheet!ref patterns
    plain_pattern = r"(?<![!\w])\$?([A-Z]{1,3})\$?(\d+)(?::\$?([A-Z]{1,3})\$?(\d+))?"
    for m in re.finditer(plain_pattern, formula_str):
        col_start, row_start = m.group(1), int(m.group(2))
        if m.group(3) and m.group(4):
            col_end, row_end = m.group(3), int(m.group(4))
            for r in range(row_start, row_end + 1):
                for c_idx in range(_col_to_num(col_start), _col_to_num(col_end) + 1):
                    refs.add((sheet_name, f"{get_column_letter(c_idx)}{r}"))
        else:
            refs.add((sheet_name, f"{col_start}{row_start}"))

    return refs


def _col_to_num(col_str):
    """Convert column letter(s) to number (A=1, B=2, ..., AA=27)."""
    num = 0
    for c in col_str.upper():
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def inspect_workbook(filepath):
    """Analyse an Excel workbook and return a structured dict."""
    wb = load_workbook(filepath)

    result = {
        'file': filepath,
        'sheets': wb.sheetnames,
        'sheet_details': {},
        'named_ranges': [],
        'dependency_analysis': {
            'potential_inputs': [],
            'potential_outputs': [],
            'intermediates': [],
        },
    }

    # Collect all cells with values and formulas for dependency analysis
    all_cells = {}  # (sheet, coord) -> value
    formula_cells = {}  # (sheet, coord) -> formula string
    dependents = defaultdict(set)  # cell -> set of cells that reference it
    precedents = defaultdict(set)  # cell -> set of cells it references

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_info = {
            'dimensions': ws.dimensions,
            'cells': [],
            'conditional_formatting': [],
            'data_validations': [],
            'charts': [],
        }

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                coord = cell.coordinate
                key = (sheet_name, coord)
                all_cells[key] = cell.value

                cell_entry = {
                    'coord': coord,
                    'value': cell.value,
                }

                color = get_cell_color(cell)
                if color:
                    cell_entry['fill'] = color

                if str(cell.value).startswith('='):
                    cell_entry['formula'] = str(cell.value)
                    formula_cells[key] = str(cell.value)

                    dynamic = detect_dynamic_refs(str(cell.value))
                    if dynamic:
                        cell_entry['dynamic_refs'] = dynamic

                if cell.comment:
                    cell_entry['comment'] = cell.comment.text

                sheet_info['cells'].append(cell_entry)

        if ws.conditional_formatting:
            for cf_rule in ws.conditional_formatting:
                for rule in cf_rule.rules:
                    entry = {
                        'range': str(cf_rule.cells),
                        'type': rule.type,
                    }
                    if rule.formula:
                        entry['formula'] = list(rule.formula)
                    elif rule.operator:
                        entry['operator'] = rule.operator
                    sheet_info['conditional_formatting'].append(entry)

        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                sheet_info['data_validations'].append({
                    'range': str(dv.sqref),
                    'type': dv.type or 'list',
                    'formula1': dv.formula1 or '',
                })

        if ws._charts:
            for i, chart in enumerate(ws._charts):
                chart_entry = {
                    'index': i + 1,
                    'type': type(chart).__name__,
                    'title': chart.title or None,
                }
                if hasattr(chart, 'x_axis') and chart.x_axis and chart.x_axis.title:
                    chart_entry['x_axis'] = str(chart.x_axis.title)
                if hasattr(chart, 'y_axis') and chart.y_axis and chart.y_axis.title:
                    chart_entry['y_axis'] = str(chart.y_axis.title)
                series_list = []
                if hasattr(chart, 'series'):
                    for j, series in enumerate(chart.series):
                        data_ref = series.val.numRef.f if series.val and series.val.numRef else None
                        series_list.append({
                            'index': j + 1,
                            'data': data_ref,
                        })
                if series_list:
                    chart_entry['series'] = series_list
                sheet_info['charts'].append(chart_entry)

        result['sheet_details'][sheet_name] = sheet_info

    # Named ranges
    if wb.defined_names:
        for dn in wb.defined_names:
            if hasattr(dn, 'name'):
                result['named_ranges'].append({
                    'name': dn.name,
                    'reference': dn.attr_text,
                })
            else:
                # Some openpyxl versions yield strings instead of
                # DefinedName objects when iterating DefinedNameList.
                entry = wb.defined_names[dn]
                result['named_ranges'].append({
                    'name': entry.name if hasattr(entry, 'name') else str(dn),
                    'reference': entry.attr_text if hasattr(entry, 'attr_text') else '',
                })

    # Dependency analysis
    for (sheet, coord), formula in formula_cells.items():
        refs = parse_cell_refs(formula, sheet)
        for ref in refs:
            precedents[(sheet, coord)].add(ref)
            dependents[ref].add((sheet, coord))

    valued_cells = set(all_cells.keys())

    for key in valued_cells:
        if key not in formula_cells and key in dependents:
            sheet, coord = key
            result['dependency_analysis']['potential_inputs'].append({
                'sheet': sheet,
                'coord': coord,
                'value': all_cells[key],
                'dependent_count': len(dependents[key]),
            })

    for key in formula_cells:
        if key not in dependents:
            sheet, coord = key
            result['dependency_analysis']['potential_outputs'].append({
                'sheet': sheet,
                'coord': coord,
                'formula': formula_cells[key],
            })

    for key in formula_cells:
        if key in dependents:
            sheet, coord = key
            result['dependency_analysis']['intermediates'].append({
                'sheet': sheet,
                'coord': coord,
                'formula': formula_cells[key],
            })

    return result


def print_text(data):
    """Print workbook analysis in human-readable text format."""
    print(f"=== Workbook: {data['file']} ===")
    print(f"Sheets: {', '.join(data['sheets'])}")
    print()

    for sheet_name, info in data['sheet_details'].items():
        print(f"--- Sheet: {sheet_name} ---")
        print(f"Dimensions: {info['dimensions']}")
        print()

        if info['cells']:
            print("Cells:")
            for c in info['cells']:
                parts = [f"  {c['coord']}: {c['value']}"]
                if 'fill' in c:
                    parts.append(f"[fill: {c['fill']}]")
                if 'formula' in c:
                    parts.append("(formula)")
                if 'dynamic_refs' in c:
                    parts.append(f"[dynamic: {', '.join(c['dynamic_refs'])}]")
                if 'comment' in c:
                    parts.append(f"[comment: {c['comment']}]")
                print("  ".join(parts))
            print()

        if info['conditional_formatting']:
            print("Conditional formatting:")
            for cf in info['conditional_formatting']:
                desc = ""
                if 'formula' in cf:
                    desc = f" formula={cf['formula']}"
                elif 'operator' in cf:
                    desc = f" operator={cf['operator']}"
                print(f"  Range {cf['range']}: type={cf['type']}{desc}")
            print()

        if info['data_validations']:
            print("Data validations:")
            for dv in info['data_validations']:
                print(f"  Range {dv['range']}: type={dv['type']} formula1={dv['formula1']}")
            print()

        if info['charts']:
            print("Charts:")
            for ch in info['charts']:
                print(f"  Chart {ch['index']}: type={ch['type']}, title={ch.get('title', '(no title)')}")
                if 'x_axis' in ch:
                    print(f"    X-axis: {ch['x_axis']}")
                if 'y_axis' in ch:
                    print(f"    Y-axis: {ch['y_axis']}")
                if 'series' in ch:
                    for s in ch['series']:
                        print(f"    Series {s['index']}: data={s.get('data', '?')}")
            print()

    if data['named_ranges']:
        print("--- Named Ranges ---")
        for nr in data['named_ranges']:
            print(f"  {nr['name']}: {nr['reference']}")
        print()

    dep = data['dependency_analysis']
    print("--- Formula Dependency Analysis ---")
    print()

    if dep['potential_inputs']:
        print("Potential INPUTS (non-formula cells that other formulas depend on):")
        for item in dep['potential_inputs']:
            print(f"  {item['sheet']}!{item['coord']}: {item['value']}  ({item['dependent_count']} dependents)")
        print()

    if dep['potential_outputs']:
        print("Potential OUTPUTS (formula cells that nothing else references):")
        for item in dep['potential_outputs']:
            print(f"  {item['sheet']}!{item['coord']}: {item['formula']}")
        print()

    if dep['intermediates']:
        print("Intermediate calculations (formula cells referenced by other formulas):")
        for item in dep['intermediates']:
            print(f"  {item['sheet']}!{item['coord']}: {item['formula']}")
        print()

    if not dep['potential_inputs'] and not dep['potential_outputs']:
        print("No formula dependencies found.")
        print()


def main():
    parser = argparse.ArgumentParser(
        description='Inspect an Excel file to plan its conversion into a Molnify app.')
    parser.add_argument('file', help='Excel file to inspect (.xlsx)')
    parser.add_argument('--json', action='store_true',
                        help='Output raw JSON instead of the default human-readable text')
    args = parser.parse_args()

    data = inspect_workbook(args.file)

    if args.json:
        print(json.dumps(data, indent=2, default=str))
    else:
        print_text(data)


if __name__ == '__main__':
    main()
